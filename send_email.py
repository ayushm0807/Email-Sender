import smtplib
import os
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv  # Load environment variables

# Load environment variables from .env file
load_dotenv()

# Get credentials from environment variables
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))  # Default to 587 if not found

# Load the Excel sheet
file_path = "/home/ayushmishra/Felix/Email Sender/email_list.xlsx"

# Open the existing Excel file to preserve formatting
wb = load_workbook(file_path)
ws = wb.active

# Process each row in the Excel file
for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
    to_email = ws.cell(row=row, column=3).value  # Email is in column C (3)
    send_value = ws.cell(row=row, column=8).value  # "Send" is in column H (8)
    pdf_path = ws.cell(row=row, column=7).value  # "PDF Path" is in column G (7)

    # Skip completely empty rows
    if not any(ws.cell(row=row, column=col).value for col in range(1, 10)):  
        continue  # If all cells in the row are empty, skip it

    # If "Send" is not "Yes", then it is not "Sent"
    if str(send_value).strip().lower() != "yes":
        ws.cell(row=row, column=9, value="No")  # Update "Sent" column (I) to "No"
        continue

    # Skip if already sent
    sent_value = ws.cell(row=row, column=9).value  # "Sent" is in column I (9)
    if str(sent_value).strip().lower() == "yes":
        continue

    name = ws.cell(row=row, column=2).value  # "Name" is in column B (2)
    company = ws.cell(row=row, column=4).value  # "Company" is in column D (4)
    subject = ws.cell(row=row, column=5).value  # "Subject" is in column E (5)
    description = ws.cell(row=row, column=6).value  # "Description" is in column F (6)

    # Replace placeholders with actual values
    description = description.replace("{name}", name).replace("{company}", company)

    try:
        # Create Email
        msg = MIMEMultipart()
        msg["From"] = EMAIL_SENDER
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(description, "plain"))

        # Ensure `pdf_path` is a valid string
        if pdf_path:
            pdf_path = pdf_path.replace("\u200b", "").strip()  # Remove invisible characters, not spaces

        # Attach PDF file if it exists
        if pdf_path and os.path.exists(pdf_path):
            with open(pdf_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())

            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(pdf_path)}"')
            msg.attach(part)
            print(f"Attached PDF: {os.path.basename(pdf_path)}")

        # Send Email
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(msg["From"], to_email, msg.as_string())

        print(f"Email sent to {to_email}")
        
        # Update "Sent" column (I) to "Yes" after sending the email
        ws.cell(row=row, column=9, value="Yes")

    except Exception as e:
        print(f"Failed to send email to {to_email}: {e}")
        
        # Update "Sent" column (I) to "Failed" if sending fails
        ws.cell(row=row, column=9, value="Failed")

# Save the workbook with updated data
wb.save(file_path)
