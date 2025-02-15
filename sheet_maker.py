import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# Define file path
file_path = "/home/ayushmishra/Felix/Email Sender/email_list.xlsx"

# Load the existing workbook
wb = load_workbook(file_path)
ws = wb.active

# Ensure the first row contains correct headers
headers = ["S.No.", "Name", "Mail", "Company", "Subject", "Description", "PDF Path", "Send", "Sent"]

# If the first row is empty, add headers
if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
    ws.append(headers)

# Set the column widths
column_widths = {
    'A': 5,   # S.No.
    'B': 20,  # Name
    'C': 25,  # Mail
    'D': 20,  # Company
    'E': 40,  # Subject
    'F': 80,  # Description
    'G': 20,  # PDF Path (Ensure this column exists)
    'H': 10,  # Send
    'I': 10   # Sent
}

# Apply column widths
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Apply Conditional Formatting to "Sent" column (Column I)
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for "Yes"
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red for "No"

ws.conditional_formatting.add('I2:I1000', FormulaRule(formula=['$I2="Yes"'], fill=green_fill))
ws.conditional_formatting.add('I2:I1000', FormulaRule(formula=['$I2="No"'], fill=red_fill))

# Increase the row height for the first 1000 rows
for row in range(1, 1001):  # Rows 1 to 1000
    ws.row_dimensions[row].height = 30  # Set row height

# Save the updated workbook
wb.save(file_path)

file_path  # Return the file path for confirmation
